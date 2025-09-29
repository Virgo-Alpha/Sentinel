"""
QueryKB Lambda tool for natural language query processing.

This Lambda function handles natural language query processing, DynamoDB and OpenSearch
query translation, filtering by date range, keywords, categories, and sources,
result ranking and relevance scoring for the Sentinel cybersecurity triage system.
"""

import json
import logging
import os
from datetime import datetime, timezone, timedelta
from typing import Any, Dict, List, Optional, Union
from dataclasses import dataclass
from decimal import Decimal
import re
import uuid

import boto3
from botocore.exceptions import ClientError
from boto3.dynamodb.conditions import Key, Attr

# Configure logging
logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)

# AWS clients
dynamodb = boto3.resource('dynamodb')
bedrock_runtime = boto3.client('bedrock-runtime')


@dataclass
class QueryResult:
    """Individual query result item."""
    article_id: str
    title: str
    url: str
    published_at: datetime
    keyword_matches: List[str]
    hit_count: int
    description: str
    relevancy_score: Optional[float] = None
    source: Optional[str] = None
    tags: Optional[List[str]] = None


@dataclass
class QueryResponse:
    """Complete query response."""
    success: bool
    total_results: int
    results: List[QueryResult]
    query_time_ms: int
    filters_applied: Dict[str, Any]
    errors: List[str] = None
    warnings: List[str] = None
    export_url: Optional[str] = None
    
    def __post_init__(self):
        if self.errors is None:
            self.errors = []
        if self.warnings is None:
            self.warnings = []


class QueryKBError(Exception):
    """Custom exception for QueryKB errors."""
    pass

class NaturalLanguageProcessor:
    """Processes natural language queries and extracts structured filters."""
    
    def __init__(self, bedrock_model_id: str):
        self.bedrock_model_id = bedrock_model_id
    
    def parse_query(self, query: str) -> Dict[str, Any]:
        """Parse natural language query into structured filters."""
        try:
            logger.info(f"Parsing natural language query: {query}")
            
            # Create prompt for query parsing
            prompt = self._create_query_parsing_prompt(query)
            
            # Call Bedrock to parse the query
            response = bedrock_runtime.invoke_model(
                modelId=self.bedrock_model_id,
                body=json.dumps({
                    "anthropic_version": "bedrock-2023-05-31",
                    "max_tokens": 1000,
                    "messages": [
                        {
                            "role": "user",
                            "content": prompt
                        }
                    ]
                })
            )
            
            response_body = json.loads(response['body'].read())
            parsed_content = response_body['content'][0]['text']
            
            # Extract JSON from the response
            try:
                # Find JSON in the response
                json_match = re.search(r'\{.*\}', parsed_content, re.DOTALL)
                if json_match:
                    parsed_filters = json.loads(json_match.group())
                else:
                    # Fallback to basic parsing
                    parsed_filters = self._basic_query_parsing(query)
            except json.JSONDecodeError:
                logger.warning("Failed to parse LLM response as JSON, using basic parsing")
                parsed_filters = self._basic_query_parsing(query)
            
            logger.info(f"Parsed filters: {parsed_filters}")
            return parsed_filters
            
        except Exception as e:
            logger.error(f"Error parsing natural language query: {e}")
            # Fallback to basic parsing
            return self._basic_query_parsing(query)
    
    def _create_query_parsing_prompt(self, query: str) -> str:
        """Create prompt for query parsing."""
        return f"""
Parse the following natural language query into structured search filters for a cybersecurity intelligence database.

Query: "{query}"

Extract the following information and return as JSON:
- keywords: List of specific keywords or terms to search for
- date_range: Object with "start" and "end" dates if mentioned (ISO format)
- categories: List of feed categories (Advisories, Alerts, Vulnerabilities, Vendor, Threat Intel, Research, News, Data Breach, Policy)
- sources: List of specific sources or feeds mentioned
- entities: Object with CVEs, threat actors, malware, vendors, products if mentioned
- intent: The main intent of the query (search, report, analysis, etc.)

Return only valid JSON without any additional text.

Example:
{{
  "keywords": ["Microsoft", "Azure", "vulnerability"],
  "date_range": {{"start": "2024-01-01T00:00:00Z", "end": "2024-01-31T23:59:59Z"}},
  "categories": ["Vulnerabilities", "Advisories"],
  "sources": [],
  "entities": {{"cves": ["CVE-2024-1234"], "vendors": ["Microsoft"]}},
  "intent": "search"
}}
"""
    
    def _basic_query_parsing(self, query: str) -> Dict[str, Any]:
        """Basic query parsing as fallback."""
        filters = {
            "keywords": [],
            "date_range": None,
            "categories": [],
            "sources": [],
            "entities": {"cves": [], "vendors": [], "products": []},
            "intent": "search"
        }
        
        # Extract keywords (simple word extraction)
        words = re.findall(r'\b\w+\b', query.lower())
        
        # Common cybersecurity terms
        cyber_terms = [
            "vulnerability", "cve", "malware", "breach", "attack", "threat",
            "security", "exploit", "patch", "advisory", "alert"
        ]
        
        # Technology keywords
        tech_terms = [
            "microsoft", "azure", "aws", "google", "cisco", "fortinet",
            "windows", "linux", "apache", "nginx", "kubernetes"
        ]
        
        # Extract relevant keywords
        for word in words:
            if word in cyber_terms or word in tech_terms or len(word) > 4:
                filters["keywords"].append(word)
        
        # Extract CVEs
        cve_pattern = r'CVE-\d{4}-\d{4,7}'
        cves = re.findall(cve_pattern, query.upper())
        filters["entities"]["cves"] = cves
        
        # Extract date references
        if any(term in query.lower() for term in ["today", "yesterday", "week", "month"]):
            now = datetime.now(timezone.utc)
            if "today" in query.lower():
                filters["date_range"] = {
                    "start": now.replace(hour=0, minute=0, second=0).isoformat(),
                    "end": now.isoformat()
                }
            elif "week" in query.lower():
                week_ago = now - timedelta(days=7)
                filters["date_range"] = {
                    "start": week_ago.isoformat(),
                    "end": now.isoformat()
                }
            elif "month" in query.lower():
                month_ago = now - timedelta(days=30)
                filters["date_range"] = {
                    "start": month_ago.isoformat(),
                    "end": now.isoformat()
                }
        
        return filters


class DatabaseQueryEngine:
    """Handles database queries against DynamoDB and OpenSearch."""
    
    def __init__(self, articles_table_name: str, opensearch_endpoint: Optional[str] = None):
        self.articles_table = dynamodb.Table(articles_table_name)
        self.opensearch_client = None
        
        if opensearch_endpoint:
            try:
                from opensearchpy import OpenSearch, RequestsHttpConnection
                from aws_requests_auth.aws_auth import AWSRequestsAuth
                
                # Initialize OpenSearch client
                region = os.environ.get('AWS_REGION', 'us-east-1')
                credentials = boto3.Session().get_credentials()
                awsauth = AWSRequestsAuth(credentials, region, 'es')
                
                self.opensearch_client = OpenSearch(
                    hosts=[{'host': opensearch_endpoint.replace('https://', ''), 'port': 443}],
                    http_auth=awsauth,
                    use_ssl=True,
                    verify_certs=True,
                    connection_class=RequestsHttpConnection
                )
            except ImportError:
                logger.warning("OpenSearch dependencies not available, using DynamoDB only")
                self.opensearch_client = None
    
    def search_articles(self, filters: Dict[str, Any], limit: int = 100) -> List[QueryResult]:
        """Search articles using the provided filters."""
        try:
            logger.info(f"Searching articles with filters: {filters}")
            
            # Try OpenSearch first if available and keywords are provided
            if self.opensearch_client and filters.get("keywords"):
                try:
                    return self._search_opensearch(filters, limit)
                except Exception as e:
                    logger.warning(f"OpenSearch query failed, falling back to DynamoDB: {e}")
            
            # Fallback to DynamoDB scan/query
            return self._search_dynamodb(filters, limit)
            
        except Exception as e:
            logger.error(f"Error searching articles: {e}")
            raise QueryKBError(f"Search failed: {str(e)}")
    
    def _search_opensearch(self, filters: Dict[str, Any], limit: int) -> List[QueryResult]:
        """Search using OpenSearch for full-text search."""
        query_body = {
            "query": {
                "bool": {
                    "must": [],
                    "filter": []
                }
            },
            "sort": [{"published_at": {"order": "desc"}}],
            "size": limit
        }
        
        # Add keyword search
        keywords = filters.get("keywords", [])
        if keywords:
            keyword_query = {
                "multi_match": {
                    "query": " ".join(keywords),
                    "fields": ["title^2", "content", "summary_short", "summary_card"],
                    "type": "best_fields",
                    "fuzziness": "AUTO"
                }
            }
            query_body["query"]["bool"]["must"].append(keyword_query)
        
        # Add filters
        self._add_opensearch_filters(query_body, filters)
        
        logger.info(f"OpenSearch query: {json.dumps(query_body, indent=2)}")
        
        response = self.opensearch_client.search(
            index="sentinel-articles",
            body=query_body
        )
        
        results = []
        for hit in response['hits']['hits']:
            source = hit['_source']
            
            # Calculate keyword matches and hit count
            keyword_matches, hit_count = self._calculate_keyword_matches(
                source, keywords
            )
            
            result = QueryResult(
                article_id=source['article_id'],
                title=source['title'],
                url=source['url'],
                published_at=datetime.fromisoformat(source['published_at'].replace('Z', '+00:00')),
                keyword_matches=keyword_matches,
                hit_count=hit_count,
                description=source.get('summary_short', source.get('title', '')),
                relevancy_score=source.get('relevancy_score'),
                source=source.get('source'),
                tags=source.get('tags', [])
            )
            results.append(result)
        
        return results
    
    def _search_dynamodb(self, filters: Dict[str, Any], limit: int) -> List[QueryResult]:
        """Search using DynamoDB scan with filters."""
        scan_kwargs = {
            'Limit': limit,
            'FilterExpression': None
        }
        
        filter_expressions = []
        
        # Add state filter (only published and processed articles)
        filter_expressions.append(
            Attr('state').is_in(['PUBLISHED', 'PROCESSED'])
        )
        
        # Add date range filter
        date_range = filters.get('date_range')
        if date_range:
            if date_range.get('start'):
                filter_expressions.append(
                    Attr('published_at').gte(date_range['start'])
                )
            if date_range.get('end'):
                filter_expressions.append(
                    Attr('published_at').lte(date_range['end'])
                )
        
        # Add category filter
        categories = filters.get('categories', [])
        if categories:
            # Map categories to feed types (this would need feed metadata)
            filter_expressions.append(
                Attr('tags').contains(categories[0])  # Simplified
            )
        
        # Add source filter
        sources = filters.get('sources', [])
        if sources:
            filter_expressions.append(
                Attr('source').is_in(sources)
            )
        
        # Combine filter expressions
        if filter_expressions:
            combined_filter = filter_expressions[0]
            for expr in filter_expressions[1:]:
                combined_filter = combined_filter & expr
            scan_kwargs['FilterExpression'] = combined_filter
        
        logger.info(f"DynamoDB scan with filters: {scan_kwargs}")
        
        response = self.articles_table.scan(**scan_kwargs)
        items = response.get('Items', [])
        
        # Handle pagination if needed
        while 'LastEvaluatedKey' in response and len(items) < limit:
            scan_kwargs['ExclusiveStartKey'] = response['LastEvaluatedKey']
            response = self.articles_table.scan(**scan_kwargs)
            items.extend(response.get('Items', []))
        
        # Convert to QueryResult objects and apply keyword filtering
        results = []
        keywords = filters.get("keywords", [])
        
        for item in items:
            # Convert DynamoDB types
            item = self._convert_from_dynamodb_types(item)
            
            # Calculate keyword matches if keywords provided
            keyword_matches, hit_count = self._calculate_keyword_matches(item, keywords)
            
            # Skip if keywords provided but no matches found
            if keywords and hit_count == 0:
                continue
            
            result = QueryResult(
                article_id=item['article_id'],
                title=item['title'],
                url=item['url'],
                published_at=datetime.fromisoformat(item['published_at'].replace('Z', '+00:00')),
                keyword_matches=keyword_matches,
                hit_count=hit_count,
                description=item.get('summary_short', item.get('title', '')),
                relevancy_score=item.get('relevancy_score'),
                source=item.get('source'),
                tags=item.get('tags', [])
            )
            results.append(result)
        
        # Sort by published date descending
        results.sort(key=lambda x: x.published_at, reverse=True)
        
        return results[:limit]
    
    def _add_opensearch_filters(self, query_body: Dict, filters: Dict[str, Any]):
        """Add filters to OpenSearch query."""
        filter_clauses = query_body["query"]["bool"]["filter"]
        
        # Date range filter
        date_range = filters.get('date_range')
        if date_range:
            range_filter = {"range": {"published_at": {}}}
            if date_range.get('start'):
                range_filter["range"]["published_at"]["gte"] = date_range['start']
            if date_range.get('end'):
                range_filter["range"]["published_at"]["lte"] = date_range['end']
            filter_clauses.append(range_filter)
        
        # Categories filter
        categories = filters.get('categories', [])
        if categories:
            filter_clauses.append({
                "terms": {"tags": categories}
            })
        
        # Sources filter
        sources = filters.get('sources', [])
        if sources:
            filter_clauses.append({
                "terms": {"source": sources}
            })
        
        # State filter
        filter_clauses.append({
            "terms": {"state": ["PUBLISHED", "PROCESSED"]}
        })
    
    def _calculate_keyword_matches(self, article: Dict, keywords: List[str]) -> tuple[List[str], int]:
        """Calculate keyword matches and hit count for an article."""
        if not keywords:
            return [], 0
        
        matches = []
        total_hits = 0
        
        # Text fields to search
        text_fields = [
            article.get('title', ''),
            article.get('summary_short', ''),
            article.get('summary_card', ''),
            ' '.join(article.get('tags', []))
        ]
        
        # Combine all text
        combined_text = ' '.join(text_fields).lower()
        
        for keyword in keywords:
            keyword_lower = keyword.lower()
            count = combined_text.count(keyword_lower)
            if count > 0:
                matches.append(keyword)
                total_hits += count
        
        # Also check existing keyword_matches from processing
        existing_matches = article.get('keyword_matches', [])
        for match in existing_matches:
            if isinstance(match, dict):
                keyword = match.get('keyword', '')
                hit_count = match.get('hit_count', 0)
                if keyword and keyword not in matches:
                    matches.append(keyword)
                    total_hits += hit_count
        
        return matches, total_hits
    
    def _convert_from_dynamodb_types(self, item: Dict[str, Any]) -> Dict[str, Any]:
        """Convert DynamoDB types back to Python types."""
        converted = {}
        for key, value in item.items():
            if isinstance(value, Decimal):
                converted[key] = float(value)
            elif isinstance(value, dict):
                converted[key] = self._convert_from_dynamodb_types(value)
            elif isinstance(value, list):
                converted[key] = [
                    self._convert_from_dynamodb_types(item) if isinstance(item, dict) 
                    else float(item) if isinstance(item, Decimal) 
                    else item 
                    for item in value
                ]
            else:
                converted[key] = value
        return converted


class ReportGenerator:
    """Generates reports and exports from query results."""
    
    def __init__(self, artifacts_bucket: str):
        self.artifacts_bucket = artifacts_bucket
        self.s3_client = boto3.client('s3')
    
    def generate_xlsx_report(self, results: List[QueryResult], 
                           filename: Optional[str] = None) -> str:
        """Generate XLSX report from query results."""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
            from io import BytesIO
            
            # Create workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Cybersecurity Intelligence Report"
            
            # Define headers
            headers = [
                "Title", "Link", "Published Date", "Keywords", 
                "Keyword Hit Count", "Description", "Source", 
                "Relevancy Score", "Tags"
            ]
            
            # Add headers with styling
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
            
            # Add data rows
            for row, result in enumerate(results, 2):
                ws.cell(row=row, column=1, value=result.title)
                ws.cell(row=row, column=2, value=result.url)
                ws.cell(row=row, column=3, value=result.published_at.strftime("%Y-%m-%d %H:%M:%S"))
                ws.cell(row=row, column=4, value=", ".join(result.keyword_matches))
                ws.cell(row=row, column=5, value=result.hit_count)
                ws.cell(row=row, column=6, value=result.description)
                ws.cell(row=row, column=7, value=result.source or "")
                ws.cell(row=row, column=8, value=result.relevancy_score or "")
                ws.cell(row=row, column=9, value=", ".join(result.tags or []))
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save to BytesIO
            excel_buffer = BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            
            # Generate filename if not provided
            if not filename:
                timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
                filename = f"cybersecurity_report_{timestamp}.xlsx"
            
            # Upload to S3
            key = f"reports/{filename}"
            self.s3_client.put_object(
                Bucket=self.artifacts_bucket,
                Key=key,
                Body=excel_buffer.getvalue(),
                ContentType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            
            # Generate presigned URL
            s3_url = self.s3_client.generate_presigned_url(
                'get_object',
                Params={'Bucket': self.artifacts_bucket, 'Key': key},
                ExpiresIn=3600  # 1 hour
            )
            
            logger.info(f"Generated XLSX report: {key}")
            return s3_url
            
        except ImportError:
            logger.error("openpyxl not available, cannot generate XLSX report")
            raise QueryKBError("XLSX generation not available - openpyxl not installed")
        except Exception as e:
            logger.error(f"Error generating XLSX report: {e}")
            raise QueryKBError(f"Report generation failed: {str(e)}")


class QueryKBTool:
    """Main QueryKB tool orchestrating natural language processing and database queries."""
    
    def __init__(self, articles_table_name: str, opensearch_endpoint: Optional[str],
                 artifacts_bucket: str, bedrock_model_id: str):
        self.nlp_processor = NaturalLanguageProcessor(bedrock_model_id)
        self.query_engine = DatabaseQueryEngine(articles_table_name, opensearch_endpoint)
        self.report_generator = ReportGenerator(artifacts_bucket)
    
    def process_query(self, query: str, filters: Optional[Dict[str, Any]] = None,
                     export_format: Optional[str] = None, limit: int = 100) -> QueryResponse:
        """Process a natural language query and return results."""
        start_time = datetime.now()
        
        try:
            logger.info(f"Processing query: {query}")
            
            # Parse natural language query if no explicit filters provided
            if not filters:
                filters = self.nlp_processor.parse_query(query)
            
            # Search articles
            results = self.query_engine.search_articles(filters, limit)
            
            # Generate export if requested
            export_url = None
            if export_format == "xlsx" and results:
                export_url = self.report_generator.generate_xlsx_report(results)
            
            # Calculate query time
            query_time_ms = int((datetime.now() - start_time).total_seconds() * 1000)
            
            return QueryResponse(
                success=True,
                total_results=len(results),
                results=results,
                query_time_ms=query_time_ms,
                filters_applied=filters,
                export_url=export_url
            )
            
        except Exception as e:
            logger.error(f"Error processing query: {e}")
            query_time_ms = int((datetime.now() - start_time).total_seconds() * 1000)
            
            return QueryResponse(
                success=False,
                total_results=0,
                results=[],
                query_time_ms=query_time_ms,
                filters_applied=filters or {},
                errors=[str(e)]
            )


def lambda_handler(event: Dict[str, Any], context: Any) -> Dict[str, Any]:
    """Lambda handler for QueryKB operations."""
    try:
        # Extract parameters
        query = event.get('query', '')
        filters = event.get('filters')
        export_format = event.get('export_format')
        limit = event.get('limit', 100)
        
        if not query and not filters:
            raise ValueError("Either 'query' or 'filters' must be provided")
        
        # Get configuration from environment
        articles_table = os.environ.get('ARTICLES_TABLE', 'sentinel-articles')
        opensearch_endpoint = os.environ.get('OPENSEARCH_ENDPOINT')
        artifacts_bucket = os.environ.get('ARTIFACTS_BUCKET', 'sentinel-artifacts')
        bedrock_model_id = os.environ.get('BEDROCK_MODEL_ID', 'anthropic.claude-3-5-sonnet-20241022-v2:0')
        
        # Initialize QueryKB tool
        query_kb = QueryKBTool(
            articles_table, opensearch_endpoint, artifacts_bucket, bedrock_model_id
        )
        
        # Process query
        response = query_kb.process_query(query, filters, export_format, limit)
        
        # Format results for JSON serialization
        formatted_results = []
        for result in response.results:
            formatted_results.append({
                'article_id': result.article_id,
                'title': result.title,
                'url': result.url,
                'published_at': result.published_at.isoformat(),
                'keyword_matches': result.keyword_matches,
                'hit_count': result.hit_count,
                'description': result.description,
                'relevancy_score': result.relevancy_score,
                'source': result.source,
                'tags': result.tags
            })
        
        return {
            'statusCode': 200 if response.success else 400,
            'body': {
                'success': response.success,
                'total_results': response.total_results,
                'results': formatted_results,
                'query_time_ms': response.query_time_ms,
                'filters_applied': response.filters_applied,
                'errors': response.errors,
                'warnings': response.warnings,
                'export_url': response.export_url
            }
        }
        
    except Exception as e:
        logger.error(f"QueryKB operation failed: {e}")
        return {
            'statusCode': 500,
            'body': {
                'success': False,
                'error': str(e),
                'error_type': type(e).__name__
            }
        }


# For testing
if __name__ == "__main__":
    # Test event
    test_event = {
        "query": "Show me Microsoft Azure vulnerabilities from the last week",
        "export_format": "xlsx",
        "limit": 50
    }
    
    import os
    os.environ.update({
        'ARTICLES_TABLE': 'test-articles',
        'OPENSEARCH_ENDPOINT': 'https://test-opensearch.us-east-1.es.amazonaws.com',
        'ARTIFACTS_BUCKET': 'test-artifacts',
        'BEDROCK_MODEL_ID': 'anthropic.claude-3-5-sonnet-20241022-v2:0'
    })
    
    result = lambda_handler(test_event, None)
    print(json.dumps(result, indent=2))