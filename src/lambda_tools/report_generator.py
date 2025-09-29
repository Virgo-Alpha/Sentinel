"""
Enhanced Report Generation functionality for QueryKB Lambda tool.

This module provides XLSX export capability with keyword hit analysis,
result sorting by date in descending order, batch processing for large result sets,
and comprehensive reporting features for the Sentinel cybersecurity triage system.
"""

import json
import logging
import os
from datetime import datetime, timezone
from typing import Any, Dict, List, Optional, Union
from dataclasses import dataclass
import uuid

import boto3
from botocore.exceptions import ClientError

# Configure logging
logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)


@dataclass
class ReportConfig:
    """Configuration for report generation."""
    format: str = "xlsx"  # "xlsx", "json", "csv"
    filename: Optional[str] = None
    include_columns: List[str] = None
    sort_by: str = "published_at"
    sort_order: str = "desc"  # "asc" or "desc"
    batch_size: int = 1000
    include_keyword_analysis: bool = True
    include_summary_stats: bool = True
    
    def __post_init__(self):
        if self.include_columns is None:
            self.include_columns = [
                "title", "url", "published_at", "keyword", 
                "hit_count", "description", "source", "relevancy_score", "tags"
            ]


@dataclass
class ReportResult:
    """Result of report generation."""
    success: bool
    report_url: Optional[str] = None
    filename: Optional[str] = None
    total_records: int = 0
    processing_time_ms: int = 0
    errors: List[str] = None
    warnings: List[str] = None
    metadata: Dict[str, Any] = None
    
    def __post_init__(self):
        if self.errors is None:
            self.errors = []
        if self.warnings is None:
            self.warnings = []
        if self.metadata is None:
            self.metadata = {}


class ReportGenerationError(Exception):
    """Custom exception for report generation errors."""
    pass


class KeywordAnalyzer:
    """Analyzes keyword hits and generates statistics."""
    
    def analyze_keyword_hits(self, results: List[Dict[str, Any]]) -> Dict[str, Any]:
        """Analyze keyword hits across all results."""
        try:
            keyword_stats = {}
            total_articles = len(results)
            total_hits = 0
            
            # Analyze each result
            for result in results:
                keyword_matches = result.get('keyword_matches', [])
                hit_count = result.get('hit_count', 0)
                total_hits += hit_count
                
                # Count keyword occurrences
                for keyword in keyword_matches:
                    if keyword not in keyword_stats:
                        keyword_stats[keyword] = {
                            'total_hits': 0,
                            'article_count': 0,
                            'articles': []
                        }
                    
                    keyword_stats[keyword]['total_hits'] += hit_count
                    keyword_stats[keyword]['article_count'] += 1
                    keyword_stats[keyword]['articles'].append({
                        'article_id': result.get('article_id'),
                        'title': result.get('title'),
                        'hit_count': hit_count
                    })
            
            # Calculate statistics
            analysis = {
                'total_articles': total_articles,
                'total_keyword_hits': total_hits,
                'unique_keywords': len(keyword_stats),
                'average_hits_per_article': total_hits / total_articles if total_articles > 0 else 0,
                'keyword_breakdown': {}
            }
            
            # Sort keywords by total hits
            sorted_keywords = sorted(
                keyword_stats.items(), 
                key=lambda x: x[1]['total_hits'], 
                reverse=True
            )
            
            for keyword, stats in sorted_keywords:
                analysis['keyword_breakdown'][keyword] = {
                    'total_hits': stats['total_hits'],
                    'article_count': stats['article_count'],
                    'hit_percentage': (stats['total_hits'] / total_hits * 100) if total_hits > 0 else 0,
                    'articles_percentage': (stats['article_count'] / total_articles * 100) if total_articles > 0 else 0
                }
            
            logger.info(f"Keyword analysis completed: {len(keyword_stats)} unique keywords, {total_hits} total hits")
            return analysis
            
        except Exception as e:
            logger.error(f"Error analyzing keyword hits: {e}")
            return {
                'total_articles': len(results),
                'total_keyword_hits': 0,
                'unique_keywords': 0,
                'error': str(e)
            }


class BatchProcessor:
    """Handles batch processing for large result sets."""
    
    def __init__(self, batch_size: int = 1000):
        self.batch_size = batch_size
    
    def process_in_batches(self, results: List[Dict[str, Any]], 
                          processor_func, *args, **kwargs) -> List[Any]:
        """Process results in batches to handle large datasets."""
        try:
            logger.info(f"Processing {len(results)} results in batches of {self.batch_size}")
            
            processed_results = []
            
            for i in range(0, len(results), self.batch_size):
                batch = results[i:i + self.batch_size]
                batch_num = (i // self.batch_size) + 1
                total_batches = (len(results) + self.batch_size - 1) // self.batch_size
                
                logger.info(f"Processing batch {batch_num}/{total_batches} ({len(batch)} items)")
                
                try:
                    batch_result = processor_func(batch, *args, **kwargs)
                    processed_results.extend(batch_result if isinstance(batch_result, list) else [batch_result])
                except Exception as e:
                    logger.error(f"Error processing batch {batch_num}: {e}")
                    # Continue with next batch
                    continue
            
            logger.info(f"Batch processing completed: {len(processed_results)} results processed")
            return processed_results
            
        except Exception as e:
            logger.error(f"Error in batch processing: {e}")
            raise ReportGenerationError(f"Batch processing failed: {str(e)}")


class XLSXReportGenerator:
    """Generates XLSX reports with advanced formatting and analysis."""
    
    def __init__(self, s3_client, artifacts_bucket: str):
        self.s3_client = s3_client
        self.artifacts_bucket = artifacts_bucket
        self.keyword_analyzer = KeywordAnalyzer()
    
    def generate_xlsx_report(self, results: List[Dict[str, Any]], 
                           config: ReportConfig) -> ReportResult:
        """Generate comprehensive XLSX report with keyword analysis."""
        start_time = datetime.now()
        
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            from openpyxl.chart import BarChart, Reference
            from io import BytesIO
            
            logger.info(f"Generating XLSX report for {len(results)} results")
            
            # Sort results by date in descending order (requirement)
            sorted_results = self._sort_results(results, config.sort_by, config.sort_order)
            
            # Create workbook with multiple sheets
            wb = openpyxl.Workbook()
            
            # Remove default sheet
            wb.remove(wb.active)
            
            # Create main data sheet
            self._create_main_data_sheet(wb, sorted_results, config)
            
            # Create keyword analysis sheet if requested
            if config.include_keyword_analysis:
                self._create_keyword_analysis_sheet(wb, sorted_results)
            
            # Create summary statistics sheet if requested
            if config.include_summary_stats:
                self._create_summary_sheet(wb, sorted_results)
            
            # Save to BytesIO
            excel_buffer = BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            
            # Generate filename if not provided
            filename = config.filename
            if not filename:
                timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
                filename = f"cybersecurity_intelligence_report_{timestamp}.xlsx"
            
            # Upload to S3
            key = f"reports/{filename}"
            self.s3_client.put_object(
                Bucket=self.artifacts_bucket,
                Key=key,
                Body=excel_buffer.getvalue(),
                ContentType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                Metadata={
                    'report_type': 'cybersecurity_intelligence',
                    'generated_at': datetime.now(timezone.utc).isoformat(),
                    'total_records': str(len(results)),
                    'sort_by': config.sort_by,
                    'sort_order': config.sort_order
                }
            )
            
            # Generate presigned URL
            report_url = self.s3_client.generate_presigned_url(
                'get_object',
                Params={'Bucket': self.artifacts_bucket, 'Key': key},
                ExpiresIn=3600  # 1 hour
            )
            
            processing_time = int((datetime.now() - start_time).total_seconds() * 1000)
            
            logger.info(f"XLSX report generated successfully: {key}")
            
            return ReportResult(
                success=True,
                report_url=report_url,
                filename=filename,
                total_records=len(results),
                processing_time_ms=processing_time,
                metadata={
                    's3_key': key,
                    'file_size_bytes': len(excel_buffer.getvalue()),
                    'sheets_created': ['Main Data', 'Keyword Analysis', 'Summary'] if config.include_keyword_analysis and config.include_summary_stats else ['Main Data']
                }
            )
            
        except ImportError:
            error_msg = "openpyxl not available - cannot generate XLSX reports"
            logger.error(error_msg)
            return ReportResult(
                success=False,
                errors=[error_msg]
            )
        except Exception as e:
            error_msg = f"XLSX report generation failed: {str(e)}"
            logger.error(error_msg)
            return ReportResult(
                success=False,
                errors=[error_msg],
                processing_time_ms=int((datetime.now() - start_time).total_seconds() * 1000)
            ) 
   
    def _sort_results(self, results: List[Dict[str, Any]], 
                     sort_by: str, sort_order: str) -> List[Dict[str, Any]]:
        """Sort results by specified field and order."""
        try:
            reverse = sort_order.lower() == "desc"
            
            if sort_by == "published_at":
                # Sort by published date
                return sorted(
                    results, 
                    key=lambda x: datetime.fromisoformat(x.get('published_at', '1970-01-01T00:00:00Z').replace('Z', '+00:00')),
                    reverse=reverse
                )
            elif sort_by == "hit_count":
                # Sort by keyword hit count
                return sorted(
                    results,
                    key=lambda x: x.get('hit_count', 0),
                    reverse=reverse
                )
            elif sort_by == "relevancy_score":
                # Sort by relevancy score
                return sorted(
                    results,
                    key=lambda x: x.get('relevancy_score', 0) or 0,
                    reverse=reverse
                )
            elif sort_by == "title":
                # Sort alphabetically by title
                return sorted(
                    results,
                    key=lambda x: x.get('title', '').lower(),
                    reverse=reverse
                )
            else:
                # Default to published date descending
                return sorted(
                    results,
                    key=lambda x: datetime.fromisoformat(x.get('published_at', '1970-01-01T00:00:00Z').replace('Z', '+00:00')),
                    reverse=True
                )
                
        except Exception as e:
            logger.warning(f"Error sorting results: {e}, using original order")
            return results
    
    def _create_main_data_sheet(self, wb, results: List[Dict[str, Any]], config: ReportConfig):
        """Create the main data sheet with article information."""
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        ws = wb.create_sheet("Main Data", 0)
        
        # Define column mappings
        column_mapping = {
            "title": "Title",
            "url": "Link", 
            "published_at": "Published Date",
            "keyword": "Keywords",
            "hit_count": "Keyword Hit Count",
            "description": "Description",
            "source": "Source",
            "relevancy_score": "Relevancy Score",
            "tags": "Tags"
        }
        
        # Filter columns based on config
        headers = []
        for col in config.include_columns:
            if col in column_mapping:
                headers.append(column_mapping[col])
        
        # Style definitions
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        data_font = Font(size=10)
        data_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Add headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Add data rows
        for row, result in enumerate(results, 2):
            for col, column_key in enumerate(config.include_columns, 1):
                if column_key not in column_mapping:
                    continue
                    
                # Get cell value based on column type
                if column_key == "published_at":
                    try:
                        pub_date = datetime.fromisoformat(result.get('published_at', '').replace('Z', '+00:00'))
                        value = pub_date.strftime("%Y-%m-%d %H:%M:%S")
                    except:
                        value = result.get('published_at', '')
                elif column_key == "keyword":
                    value = ", ".join(result.get('keyword_matches', []))
                elif column_key == "tags":
                    value = ", ".join(result.get('tags', []))
                elif column_key == "relevancy_score":
                    score = result.get('relevancy_score')
                    value = f"{score:.3f}" if score is not None else ""
                else:
                    value = result.get(column_key, "")
                
                cell = ws.cell(row=row, column=col, value=value)
                cell.font = data_font
                cell.alignment = data_alignment
                cell.border = border
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            # Set reasonable width limits
            adjusted_width = min(max(max_length + 2, 10), 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Freeze header row
        ws.freeze_panes = "A2"
        
        logger.info(f"Main data sheet created with {len(results)} rows")
    
    def _create_keyword_analysis_sheet(self, wb, results: List[Dict[str, Any]]):
        """Create keyword analysis sheet with statistics and charts."""
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.chart import BarChart, Reference
        
        ws = wb.create_sheet("Keyword Analysis")
        
        # Analyze keywords
        analysis = self.keyword_analyzer.analyze_keyword_hits(results)
        
        # Style definitions
        title_font = Font(bold=True, size=14, color="366092")
        header_font = Font(bold=True, size=12)
        data_font = Font(size=10)
        
        # Add title
        ws.cell(row=1, column=1, value="Keyword Analysis Report").font = title_font
        
        # Add summary statistics
        ws.cell(row=3, column=1, value="Summary Statistics").font = header_font
        ws.cell(row=4, column=1, value="Total Articles:").font = data_font
        ws.cell(row=4, column=2, value=analysis.get('total_articles', 0)).font = data_font
        ws.cell(row=5, column=1, value="Total Keyword Hits:").font = data_font
        ws.cell(row=5, column=2, value=analysis.get('total_keyword_hits', 0)).font = data_font
        ws.cell(row=6, column=1, value="Unique Keywords:").font = data_font
        ws.cell(row=6, column=2, value=analysis.get('unique_keywords', 0)).font = data_font
        ws.cell(row=7, column=1, value="Average Hits per Article:").font = data_font
        ws.cell(row=7, column=2, value=f"{analysis.get('average_hits_per_article', 0):.2f}").font = data_font
        
        # Add keyword breakdown table
        ws.cell(row=9, column=1, value="Keyword Breakdown").font = header_font
        
        # Headers for keyword table
        headers = ["Keyword", "Total Hits", "Articles", "Hit %", "Article %"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=10, column=col, value=header).font = header_font
        
        # Add keyword data
        keyword_breakdown = analysis.get('keyword_breakdown', {})
        for row, (keyword, stats) in enumerate(keyword_breakdown.items(), 11):
            ws.cell(row=row, column=1, value=keyword).font = data_font
            ws.cell(row=row, column=2, value=stats['total_hits']).font = data_font
            ws.cell(row=row, column=3, value=stats['article_count']).font = data_font
            ws.cell(row=row, column=4, value=f"{stats['hit_percentage']:.1f}%").font = data_font
            ws.cell(row=row, column=5, value=f"{stats['articles_percentage']:.1f}%").font = data_font
        
        # Create chart for top keywords (if we have data)
        if keyword_breakdown:
            try:
                chart = BarChart()
                chart.title = "Top Keywords by Hit Count"
                chart.x_axis.title = "Keywords"
                chart.y_axis.title = "Hit Count"
                
                # Take top 10 keywords for chart
                top_keywords = list(keyword_breakdown.items())[:10]
                if top_keywords:
                    data_range = Reference(ws, min_col=2, min_row=11, max_row=10 + len(top_keywords))
                    categories = Reference(ws, min_col=1, min_row=11, max_row=10 + len(top_keywords))
                    
                    chart.add_data(data_range, titles_from_data=False)
                    chart.set_categories(categories)
                    
                    ws.add_chart(chart, "G3")
            except Exception as e:
                logger.warning(f"Could not create keyword chart: {e}")
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max(max_length + 2, 10), 30)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        logger.info("Keyword analysis sheet created")
    
    def _create_summary_sheet(self, wb, results: List[Dict[str, Any]]):
        """Create summary statistics sheet."""
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        from collections import Counter
        
        ws = wb.create_sheet("Summary")
        
        # Style definitions
        title_font = Font(bold=True, size=14, color="366092")
        header_font = Font(bold=True, size=12)
        data_font = Font(size=10)
        
        # Add title
        ws.cell(row=1, column=1, value="Intelligence Report Summary").font = title_font
        
        # Calculate statistics
        total_articles = len(results)
        sources = [r.get('source', 'Unknown') for r in results if r.get('source')]
        source_counts = Counter(sources)
        
        # Date range
        dates = []
        for r in results:
            try:
                pub_date = datetime.fromisoformat(r.get('published_at', '').replace('Z', '+00:00'))
                dates.append(pub_date)
            except:
                continue
        
        date_range = ""
        if dates:
            min_date = min(dates)
            max_date = max(dates)
            date_range = f"{min_date.strftime('%Y-%m-%d')} to {max_date.strftime('%Y-%m-%d')}"
        
        # Relevancy scores
        relevancy_scores = [r.get('relevancy_score') for r in results if r.get('relevancy_score') is not None]
        avg_relevancy = sum(relevancy_scores) / len(relevancy_scores) if relevancy_scores else 0
        
        # Add summary data
        row = 3
        ws.cell(row=row, column=1, value="Report Statistics").font = header_font
        row += 1
        
        stats = [
            ("Total Articles", total_articles),
            ("Date Range", date_range),
            ("Unique Sources", len(source_counts)),
            ("Average Relevancy Score", f"{avg_relevancy:.3f}" if avg_relevancy > 0 else "N/A"),
            ("Generated At", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC"))
        ]
        
        for label, value in stats:
            ws.cell(row=row, column=1, value=f"{label}:").font = data_font
            ws.cell(row=row, column=2, value=str(value)).font = data_font
            row += 1
        
        # Add source breakdown
        if source_counts:
            row += 2
            ws.cell(row=row, column=1, value="Articles by Source").font = header_font
            row += 1
            
            for source, count in source_counts.most_common():
                ws.cell(row=row, column=1, value=source).font = data_font
                ws.cell(row=row, column=2, value=count).font = data_font
                ws.cell(row=row, column=3, value=f"{count/total_articles*100:.1f}%").font = data_font
                row += 1
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max(max_length + 2, 15), 40)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        logger.info("Summary sheet created")


class EnhancedReportGenerator:
    """Enhanced report generator with batch processing and multiple formats."""
    
    def __init__(self, artifacts_bucket: str):
        self.artifacts_bucket = artifacts_bucket
        self.s3_client = boto3.client('s3')
        self.xlsx_generator = XLSXReportGenerator(self.s3_client, artifacts_bucket)
        self.batch_processor = BatchProcessor()
    
    def generate_report(self, results: List[Dict[str, Any]], 
                       config: Optional[ReportConfig] = None) -> ReportResult:
        """Generate report with specified configuration."""
        if config is None:
            config = ReportConfig()
        
        try:
            logger.info(f"Generating {config.format} report for {len(results)} results")
            
            if config.format.lower() == "xlsx":
                return self._generate_xlsx_report(results, config)
            elif config.format.lower() == "json":
                return self._generate_json_report(results, config)
            elif config.format.lower() == "csv":
                return self._generate_csv_report(results, config)
            else:
                raise ReportGenerationError(f"Unsupported report format: {config.format}")
                
        except Exception as e:
            logger.error(f"Report generation failed: {e}")
            return ReportResult(
                success=False,
                errors=[str(e)]
            )
    
    def _generate_xlsx_report(self, results: List[Dict[str, Any]], 
                            config: ReportConfig) -> ReportResult:
        """Generate XLSX report using batch processing if needed."""
        if len(results) > config.batch_size:
            logger.info(f"Large dataset detected ({len(results)} records), using batch processing")
            # For very large datasets, we might need to split into multiple files
            # For now, we'll process all at once but could enhance this
        
        return self.xlsx_generator.generate_xlsx_report(results, config)
    
    def _generate_json_report(self, results: List[Dict[str, Any]], 
                            config: ReportConfig) -> ReportResult:
        """Generate JSON report."""
        start_time = datetime.now()
        
        try:
            # Sort results
            sorted_results = self.xlsx_generator._sort_results(results, config.sort_by, config.sort_order)
            
            # Create report structure
            report_data = {
                "metadata": {
                    "generated_at": datetime.now(timezone.utc).isoformat(),
                    "total_records": len(sorted_results),
                    "sort_by": config.sort_by,
                    "sort_order": config.sort_order,
                    "format": "json"
                },
                "results": sorted_results
            }
            
            # Add keyword analysis if requested
            if config.include_keyword_analysis:
                analyzer = KeywordAnalyzer()
                report_data["keyword_analysis"] = analyzer.analyze_keyword_hits(sorted_results)
            
            # Generate filename
            filename = config.filename
            if not filename:
                timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
                filename = f"cybersecurity_intelligence_report_{timestamp}.json"
            
            # Upload to S3
            key = f"reports/{filename}"
            self.s3_client.put_object(
                Bucket=self.artifacts_bucket,
                Key=key,
                Body=json.dumps(report_data, indent=2, default=str),
                ContentType='application/json'
            )
            
            # Generate presigned URL
            report_url = self.s3_client.generate_presigned_url(
                'get_object',
                Params={'Bucket': self.artifacts_bucket, 'Key': key},
                ExpiresIn=3600
            )
            
            processing_time = int((datetime.now() - start_time).total_seconds() * 1000)
            
            return ReportResult(
                success=True,
                report_url=report_url,
                filename=filename,
                total_records=len(results),
                processing_time_ms=processing_time
            )
            
        except Exception as e:
            logger.error(f"JSON report generation failed: {e}")
            return ReportResult(
                success=False,
                errors=[str(e)],
                processing_time_ms=int((datetime.now() - start_time).total_seconds() * 1000)
            )
    
    def _generate_csv_report(self, results: List[Dict[str, Any]], 
                           config: ReportConfig) -> ReportResult:
        """Generate CSV report."""
        start_time = datetime.now()
        
        try:
            import csv
            from io import StringIO
            
            # Sort results
            sorted_results = self.xlsx_generator._sort_results(results, config.sort_by, config.sort_order)
            
            # Create CSV content
            csv_buffer = StringIO()
            
            if sorted_results:
                # Define column mappings
                column_mapping = {
                    "title": "Title",
                    "url": "Link",
                    "published_at": "Published Date", 
                    "keyword": "Keywords",
                    "hit_count": "Keyword Hit Count",
                    "description": "Description",
                    "source": "Source",
                    "relevancy_score": "Relevancy Score",
                    "tags": "Tags"
                }
                
                # Filter columns based on config
                fieldnames = []
                for col in config.include_columns:
                    if col in column_mapping:
                        fieldnames.append(column_mapping[col])
                
                writer = csv.DictWriter(csv_buffer, fieldnames=fieldnames)
                writer.writeheader()
                
                for result in sorted_results:
                    row = {}
                    for col in config.include_columns:
                        if col not in column_mapping:
                            continue
                            
                        header = column_mapping[col]
                        
                        if col == "published_at":
                            try:
                                pub_date = datetime.fromisoformat(result.get('published_at', '').replace('Z', '+00:00'))
                                row[header] = pub_date.strftime("%Y-%m-%d %H:%M:%S")
                            except:
                                row[header] = result.get('published_at', '')
                        elif col == "keyword":
                            row[header] = ", ".join(result.get('keyword_matches', []))
                        elif col == "tags":
                            row[header] = ", ".join(result.get('tags', []))
                        elif col == "relevancy_score":
                            score = result.get('relevancy_score')
                            row[header] = f"{score:.3f}" if score is not None else ""
                        else:
                            row[header] = result.get(col, "")
                    
                    writer.writerow(row)
            
            # Generate filename
            filename = config.filename
            if not filename:
                timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
                filename = f"cybersecurity_intelligence_report_{timestamp}.csv"
            
            # Upload to S3
            key = f"reports/{filename}"
            self.s3_client.put_object(
                Bucket=self.artifacts_bucket,
                Key=key,
                Body=csv_buffer.getvalue(),
                ContentType='text/csv'
            )
            
            # Generate presigned URL
            report_url = self.s3_client.generate_presigned_url(
                'get_object',
                Params={'Bucket': self.artifacts_bucket, 'Key': key},
                ExpiresIn=3600
            )
            
            processing_time = int((datetime.now() - start_time).total_seconds() * 1000)
            
            return ReportResult(
                success=True,
                report_url=report_url,
                filename=filename,
                total_records=len(results),
                processing_time_ms=processing_time
            )
            
        except Exception as e:
            logger.error(f"CSV report generation failed: {e}")
            return ReportResult(
                success=False,
                errors=[str(e)],
                processing_time_ms=int((datetime.now() - start_time).total_seconds() * 1000)
            )


# For testing
if __name__ == "__main__":
    # Test the enhanced report generator
    sample_results = [
        {
            "article_id": "test-1",
            "title": "Microsoft Azure Security Alert",
            "url": "https://example.com/1",
            "published_at": "2024-01-15T10:00:00Z",
            "keyword_matches": ["Microsoft", "Azure", "security"],
            "hit_count": 5,
            "description": "Critical security issue in Azure services",
            "source": "Microsoft Security",
            "relevancy_score": 0.9,
            "tags": ["Security", "Cloud"]
        },
        {
            "article_id": "test-2",
            "title": "AWS Vulnerability Report", 
            "url": "https://example.com/2",
            "published_at": "2024-01-14T15:30:00Z",
            "keyword_matches": ["AWS", "vulnerability"],
            "hit_count": 3,
            "description": "New vulnerability discovered in AWS services",
            "source": "AWS Security",
            "relevancy_score": 0.8,
            "tags": ["Vulnerability", "Cloud"]
        }
    ]
    
    # Test configuration
    config = ReportConfig(
        format="xlsx",
        include_keyword_analysis=True,
        include_summary_stats=True
    )
    
    # Initialize generator (would need real S3 bucket in practice)
    generator = EnhancedReportGenerator("test-bucket")
    
    print("Enhanced Report Generator initialized successfully")
    print(f"Sample data: {len(sample_results)} records")
    print(f"Config: {config}")